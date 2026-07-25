"""奶茶原材料管理系统 - 后端服务
提供十大实体的通用 CRUD REST API 与静态前端托管
"""
import os
import io
import csv
import json
from fastapi import FastAPI, Depends, HTTPException, Request, File, UploadFile, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, or_, String, Integer, Float, text, inspect
from sqlalchemy.orm import Session
from urllib.parse import quote
import openpyxl

from db import engine, Base, get_db
import models

# 启动时创建表
Base.metadata.create_all(bind=engine)


def migrate():
    """为新版配料模型补齐缺失列（幂等，兼容旧库）"""
    inspector = inspect(engine)
    existing = {c["name"] for c in inspector.get_columns("ingredients")}
    new_cols = {
        "barcode": "VARCHAR(64)",
        "sale_price": "FLOAT",
        "max_stock": "FLOAT",
        "location": "VARCHAR(64)",
        "brand": "VARCHAR(64)",
        "shelf_life": "VARCHAR(32)",
        "pinyin": "VARCHAR(64)",
        "remark": "VARCHAR(256)",
        "not_count_stock": "BOOLEAN DEFAULT 0",
        "multi_code": "BOOLEAN DEFAULT 0",
        "use_member_discount": "BOOLEAN DEFAULT 0",
        "has_other_spec": "BOOLEAN DEFAULT 0",
        "label_print": "BOOLEAN DEFAULT 0",
        "refrigerated": "BOOLEAN DEFAULT 0",
        "unopened": "BOOLEAN DEFAULT 0",
        "kitchen_ticket": "BOOLEAN DEFAULT 0",
        "wholesale_price": "FLOAT DEFAULT 0",
        "prep_time": "FLOAT DEFAULT 0",
        "weight": "FLOAT DEFAULT 0",
        "min_sale_qty": "FLOAT DEFAULT 0",
        "flavor": "VARCHAR(64)",
        "tags": "VARCHAR(128)",
        "production_date": "VARCHAR(32)",
        "image_url": "VARCHAR(256)",
    }
    with engine.begin() as conn:
        for name, ddl in new_cols.items():
            if name not in existing:
                conn.execute(text(f"ALTER TABLE ingredients ADD COLUMN {name} {ddl}"))


def sync_prep_inventory(materials_json: str, store_id: int, db: Session, deduct: bool):
    """预制计划与库存同步：deduct=True 扣减用量，False 回补用量。
    优先扣减门店库存（store_id匹配），没有门店库存则扣总仓（store_id=NULL）。
    """
    try:
        items = json.loads(materials_json or "[]")
    except Exception:
        return
    if not isinstance(items, list):
        return
    for it in items:
        iid = it.get("i")
        q = float(it.get("q") or 0)
        if not iid or q <= 0:
            continue
        # 优先找门店库存，再找总仓
        inv = db.query(models.Inventory).filter(
            models.Inventory.ingredient_id == iid,
            models.Inventory.store_id == store_id
        ).first()
        if not inv:
            inv = db.query(models.Inventory).filter(
                models.Inventory.ingredient_id == iid,
                models.Inventory.store_id == None
            ).first()
        if not inv:
            continue
        inv.current_stock = max(0.0, (inv.current_stock or 0) + (-q if deduct else q))
        if deduct:
            inv.last_out = _today()
        # 按安全库存重算库存状态
        s = inv.safety_stock or 0
        if inv.current_stock <= 0:
            inv.status = "库存不足"
        elif inv.current_stock < s:
            inv.status = "低于安全库存"
        else:
            inv.status = "正常"

app = FastAPI(title="奶茶原材料管理系统", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============ 模型注册表 ============
MODEL_REGISTRY = {
    "platforms": models.Platform,
    "suppliers": models.Supplier,
    "brands": models.Brand,
    "stores": models.Store,
    "staff": models.Staff,
    "ingredients": models.Ingredient,
    "inventory": models.Inventory,
    "wastage": models.Wastage,
    "prep": models.Prep,
    "expiry": models.Expiry,
    "logistics": models.Logistics,
    "categories": models.Category,
    "units": models.Unit,
    "purchase_orders": models.PurchaseOrder,
}

# 组织层级：每个实体对其直接下级设置数量配额，由上级授权
HIERARCHY = {
    "suppliers": ("platforms", "platform_id", "max_suppliers"),
    "brands": ("suppliers", "supplier_id", "max_brands"),
    "stores": ("brands", "brand_id", "max_stores"),
    "staff": ("stores", "store_id", "max_staff"),
}


def enforce_quota(entity, data, db):
    """创建下级时校验上级配额是否已用尽（0 或不设为不限制）"""
    if entity not in HIERARCHY:
        return
    parent_entity, fk, quota_field = HIERARCHY[entity]
    pid = data.get(fk)
    if pid is None:
        return
    Parent = MODEL_REGISTRY[parent_entity]
    parent = db.query(Parent).filter(Parent.id == pid).first()
    if not parent:
        return
    quota = getattr(parent, quota_field, None)
    if quota is None or quota == 0:
        return
    Child = MODEL_REGISTRY[entity]
    used = db.query(Child).filter(getattr(Child, fk) == pid).count()
    if used >= quota:
        raise HTTPException(
            status_code=400,
            detail=f"上级「{parent.name}」的下级配额（{quota}）已用尽，无法继续添加。请在上一级记录中调整授权数量。",
        )


# ============ 权限：按组织层级做数据隔离 ============
ROLE_TO_ENTITY = {
    "supplier": "suppliers",
    "brand": "brands",
    "store": "stores",
    "staff": "staff",
}
# 上级实体 -> (下级实体列表) 用于递归收集可见子级
DOWNSTREAM = {
    "platforms": ["suppliers", "brands", "stores", "staff"],
    "suppliers": ["brands", "stores", "staff"],
    "brands": ["stores", "staff"],
    "stores": ["staff"],
}


def resolve_current_user(x_current_user: str | None, db: Session):
    """从请求头解析当前登录用户，返回 (account, entity_model, entity_obj) 或 (None,None,None)"""
    if not x_current_user:
        return None, None, None
    try:
        acct_id = int(x_current_user)
    except (ValueError, TypeError):
        return None, None, None
    acct = db.query(models.Account).filter(models.Account.id == acct_id).first()
    if not acct or not acct.role or acct.role not in ROLE_TO_ENTITY:
        return acct, None, None
    entity_name = ROLE_TO_ENTITY[acct.role]
    EntityModel = MODEL_REGISTRY[entity_name]
    obj = db.query(EntityModel).filter(EntityModel.id == acct.related_id).first()
    return acct, EntityModel, obj


def collect_visible_ids(role, entity_id, db):
    """递归收集当前用户能看到的组织实体 ID（自己 + 所有下游 + 所属上游链），返回 {entity_name: [id, ...]}"""
    visible = {}
    entity_name = ROLE_TO_ENTITY.get(role)
    if not entity_name:
        return visible
    # 自己
    visible[entity_name] = [entity_id]
    # 递归下游
    if entity_name in DOWNSTREAM:
        _collect_children(entity_name, [entity_id], visible, db)
    # 反向收集上游链（门店→品牌→供应商→平台）
    _collect_parents(entity_name, entity_id, visible, db)
    return visible


def _collect_children(parent_entity, parent_ids, visible, db):
    """递归：从 parent_entity 的 parent_ids 找所有子实体"""
    downstream = DOWNSTREAM.get(parent_entity, [])
    if not downstream:
        return
    for child_entity in downstream:
        ChildModel = MODEL_REGISTRY[child_entity]
        # 找到 parent_entity -> child_entity 的 FK 字段
        fk_col = _find_fk_to_parent(child_entity, parent_entity)
        if not fk_col:
            continue
        children = db.query(ChildModel).filter(getattr(ChildModel, fk_col).in_(parent_ids)).all()
        child_ids = [c.id for c in children]
        if child_ids:
            visible.setdefault(child_entity, []).extend(child_ids)
            _collect_children(child_entity, child_ids, visible, db)


def _collect_parents(entity_name, entity_id, visible, db):
    """反向：从当前实体往上查找所属的父实体链"""
    parent_chain = [
        ("staff", "stores", "store_id"),
        ("stores", "brands", "brand_id"),
        ("brands", "suppliers", "supplier_id"),
        ("suppliers", "platforms", "platform_id"),
    ]
    current_entity = entity_name
    current_id = entity_id
    for child_e, parent_e, fk_col in parent_chain:
        if current_entity != child_e:
            continue
        ChildModel = MODEL_REGISTRY[child_e]
        child_obj = db.query(ChildModel).filter(ChildModel.id == current_id).first()
        if not child_obj:
            break
        parent_id = getattr(child_obj, fk_col, None)
        if parent_id is None:
            break
        visible.setdefault(parent_e, []).append(parent_id)
        current_entity = parent_e
        current_id = parent_id


def _find_fk_to_parent(child_entity, parent_entity):
    """找 child 表中指向 parent 表的外键字段名"""
    hierarchy_map = {
        ("suppliers", "platforms"): "platform_id",
        ("brands", "suppliers"): "supplier_id",
        ("stores", "brands"): "brand_id",
        ("staff", "stores"): "store_id",
    }
    return hierarchy_map.get((child_entity, parent_entity))


def filter_by_visibility(q, entity, visible_ids, db):
    """对 query 施加可见性过滤：组织实体按 ID 过滤，业务实体按关联 FK 过滤"""
    # 组织实体：直接按 ID 过滤
    if entity in ("suppliers", "brands", "stores", "staff"):
        ids = visible_ids.get(entity, [])
        if not ids:
            return q.filter(text("1=0"))  # 无可见记录
        return q.filter(MODEL_REGISTRY[entity].id.in_(ids))
    # 业务实体：按关联的 FK 过滤
    if entity == "ingredients":
        # 配料：按 supplier_id 过滤
        sup_ids = visible_ids.get("suppliers", [])
        if sup_ids:
            return q.filter(models.Ingredient.supplier_id.in_(sup_ids))
    if entity == "inventory":
        ing_ids = _visible_ingredient_ids(visible_ids, db)
        if ing_ids:
            return q.filter(models.Inventory.ingredient_id.in_(ing_ids))
    if entity == "wastage":
        store_ids = visible_ids.get("stores", [])
        ing_ids = _visible_ingredient_ids(visible_ids, db)
        filters = []
        if store_ids:
            filters.append(models.Wastage.store_id.in_(store_ids))
        if ing_ids:
            filters.append(models.Wastage.ingredient_id.in_(ing_ids))
        if filters:
            return q.filter(or_(*filters))
    if entity == "prep":
        store_ids = visible_ids.get("stores", [])
        if store_ids:
            return q.filter(models.Prep.store_id.in_(store_ids))
    if entity == "expiry":
        ing_ids = _visible_ingredient_ids(visible_ids, db)
        if ing_ids:
            return q.filter(models.Expiry.ingredient_id.in_(ing_ids))
    if entity == "logistics":
        sup_ids = visible_ids.get("suppliers", [])
        store_ids = visible_ids.get("stores", [])
        filters = []
        if sup_ids:
            filters.append(models.Logistics.supplier_id.in_(sup_ids))
        if store_ids:
            filters.append(models.Logistics.store_id.in_(store_ids))
        if filters:
            return q.filter(or_(*filters))
    if entity == "purchase_orders":
        sup_ids = visible_ids.get("suppliers", [])
        store_ids = visible_ids.get("stores", [])
        filters = []
        if sup_ids:
            filters.append(models.PurchaseOrder.supplier_id.in_(sup_ids))
        if store_ids:
            filters.append(models.PurchaseOrder.store_id.in_(store_ids))
        if filters:
            # 用 AND 而非 OR：门店只看自己的订单，品牌/供应商看下游门店的订单
            return q.filter(and_(*filters))
    if entity == "inventory":
        # 库存：门店用户看门店库存+总仓；供应商看总仓
        store_ids = visible_ids.get("stores", [])
        sup_ids = visible_ids.get("suppliers", [])
        filters = []
        if store_ids:
            filters.append(or_(models.Inventory.store_id.in_(store_ids), models.Inventory.store_id == None))
        if sup_ids and not store_ids:
            filters.append(models.Inventory.store_id == None)
        if filters:
            return q.filter(or_(*filters))
    # categories, units, platforms: 不做隔离（字典/基础设施）
    return q


def _visible_ingredient_ids(visible_ids, db):
    """从可见的供应商中找出关联的配料 ID"""
    sup_ids = visible_ids.get("suppliers", [])
    if not sup_ids:
        return []
    ings = db.query(models.Ingredient.id).filter(models.Ingredient.supplier_id.in_(sup_ids)).all()
    return [i[0] for i in ings]


# 外键关联 -> 显示名称字段（list 时自动 join 填充）
NAME_JOINS = {
    "suppliers": [("platform_id", "platforms", "platform_name")],
    "brands": [("supplier_id", "suppliers", "supplier_name")],
    "stores": [("brand_id", "brands", "brand_name")],
    "staff": [("store_id", "stores", "store_name")],
    "ingredients": [("supplier_id", "suppliers", "supplier_name")],
    "inventory": [("ingredient_id", "ingredients", "ingredient_name"), ("store_id", "stores", "store_name")],
    "wastage": [("store_id", "stores", "store_name"), ("ingredient_id", "ingredients", "ingredient_name")],
    "prep": [("store_id", "stores", "store_name")],
    "expiry": [("ingredient_id", "ingredients", "ingredient_name")],
    "logistics": [("supplier_id", "suppliers", "supplier_name"), ("store_id", "stores", "store_name")],
    "purchase_orders": [("store_id", "stores", "store_name"), ("supplier_id", "suppliers", "supplier_name")],
}


def get_model(entity: str):
    if entity not in MODEL_REGISTRY:
        raise HTTPException(status_code=404, detail=f"未知实体: {entity}")
    return MODEL_REGISTRY[entity]


def get_columns(Model):
    return [c.name for c in Model.__table__.columns]


# ============ 配料管理：银豹式导入模板 / 批量导入导出 ============
INGREDIENT_HEADERS = ['名称', '分类', '条码', '规格', '主单位', '进货价', '参考售价',
                      '安全库存', '库存上限', '建议库位', '供应商', '品牌', '保质期', '拼音码', '状态', '备注']
INGREDIENT_HEADER_MAP = {
    '名称': 'name', '分类': 'category', '条码': 'barcode', '规格': 'spec', '主单位': 'unit',
    '进货价': 'cost_price', '参考售价': 'sale_price', '安全库存': 'safety_stock',
    '库存上限': 'max_stock', '建议库位': 'location', '供应商': 'supplier_id', '品牌': 'brand',
    '保质期': 'shelf_life', '拼音码': 'pinyin', '状态': 'status', '备注': 'remark',
}
INGREDIENT_FLOAT_FIELDS = {'cost_price', 'sale_price', 'safety_stock', 'max_stock'}


def _xlsx_bytes(rows):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '配料'
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _download(resp_buf, filename, ascii_name=None):
    if not ascii_name:
        ascii_name = 'download'
    ascii_name = ascii_name if ascii_name.endswith('.xlsx') else ascii_name + '.xlsx'
    disp = 'attachment; filename="%s"; filename*=UTF-8\'\'%s' % (ascii_name, quote(filename))
    return StreamingResponse(
        resp_buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': disp},
    )


def _parse_upload(content, filename):
    """解析上传的 xlsx/csv，返回 [dict(表头->值), ...]（不含表头行）"""
    if filename and filename.lower().endswith('.csv'):
        text = content.decode('utf-8-sig', errors='ignore')
        reader = list(csv.reader(io.StringIO(text)))
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        reader = [list(r) for r in wb.active.iter_rows(values_only=True)]
    if not reader:
        return []
    headers = [str(h).strip() if h is not None else '' for h in reader[0]]
    out = []
    for raw in reader[1:]:
        if raw is None or all(v is None or str(v).strip() == '' for v in raw):
            continue
        out.append({h: (raw[i] if i < len(raw) else None) for i, h in enumerate(headers)})
    return out


@app.get("/api/ingredients/template")
def ingredients_template():
    rows = [INGREDIENT_HEADERS,
            ['茉莉绿茶', '茶叶类', '690123450001', '25kg/袋', 'kg', 85, 98, 200, 500,
             'A-01-03', '茗源茶业有限公司', '鸣智优选', '180天', 'jllc', '启用', '高山茉莉绿茶，香气清幽']]
    return _download(_xlsx_bytes(rows), '配料导入模板.xlsx', 'ingredient_template')


@app.post("/api/ingredients/import")
async def ingredients_import(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    rows = _parse_upload(content, file.filename)
    created = skipped = 0
    errors = []
    suppliers = {s.name: s.id for s in db.query(models.Supplier).all()}
    for i, row in enumerate(rows, start=2):
        data = {}
        for h, v in row.items():
            field = INGREDIENT_HEADER_MAP.get(h)
            if not field or v is None:
                continue
            v = str(v).strip()
            if v == '':
                continue
            if field in INGREDIENT_FLOAT_FIELDS:
                try:
                    v = float(v)
                except (ValueError, TypeError):
                    continue
            data[field] = v
        if not data.get('name'):
            errors.append('第%d行：缺少「名称」' % i); skipped += 1; continue
        if 'supplier_id' in data:
            data['supplier_id'] = suppliers.get(data['supplier_id'])  # 名称→id，未匹配置空
        try:
            db.add(models.Ingredient(**data))
            created += 1
        except Exception as e:
            errors.append('第%d行：%s' % (i, str(e)[:80])); skipped += 1
    db.commit()
    return {'created': created, 'skipped': skipped, 'errors': errors[:20], 'total': len(rows)}


@app.get("/api/ingredients/export")
def ingredients_export(db: Session = Depends(get_db)):
    suppliers = {s.id: s.name for s in db.query(models.Supplier).all()}
    items = db.query(models.Ingredient).order_by(models.Ingredient.id).all()
    rows = [INGREDIENT_HEADERS]
    for r in items:
        rows.append([
            r.name, r.category, r.barcode, r.spec, r.unit, r.cost_price, r.sale_price,
            r.safety_stock, r.max_stock, r.location,
            suppliers.get(r.supplier_id, ''), r.brand, r.shelf_life, r.pinyin, r.status, r.remark,
        ])
    return _download(_xlsx_bytes(rows), '配料列表.xlsx', 'ingredient_list')


# ============ 通用列表（带名称 join + 搜索 + 过滤） ============
@app.get("/api/{entity}")
def list_items(entity: str, request: Request, db: Session = Depends(get_db)):
    Model = get_model(entity)
    params = dict(request.query_params)
    search = params.pop("search", None)
    page = int(params.pop("page", 1))
    page_size = int(params.pop("page_size", 200))
    filters = params

    q = db.query(Model)
    joins = NAME_JOINS.get(entity, [])
    name_aliases = []
    for fk, tgt_entity, alias in joins:
        Tgt = MODEL_REGISTRY[tgt_entity]
        q = q.join(Tgt, getattr(Model, fk) == Tgt.id, isouter=True)
        q = q.add_columns(getattr(Tgt, "name").label(alias))
        name_aliases.append(alias)

    # 精确过滤
    for f, v in filters.items():
        if f in get_columns(Model) and v not in (None, ""):
            q = q.filter(getattr(Model, f) == v)

    # 模糊搜索（所有字符串列）
    if search:
        str_cols = [getattr(Model, c) for c in get_columns(Model)
                    if isinstance(Model.__table__.c[c].type, String)]
        if str_cols:
            q = q.filter(or_(*[c.like(f"%{search}%") for c in str_cols]))

    # 组织层级数据隔离：非平台用户只能看到自己及下游的数据
    x_user = request.headers.get("x-current-user")
    acct, user_entity, user_obj = resolve_current_user(x_user, db)
    if acct and acct.role != "platform" and user_obj:
        visible_ids = collect_visible_ids(acct.role, user_obj.id, db)
        q = filter_by_visibility(q, entity, visible_ids, db)

    total = q.count()
    rows = q.offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for row in rows:
        obj = row[0] if name_aliases else row
        d = obj.to_dict()
        if name_aliases:
            for i, alias in enumerate(name_aliases):
                d[alias] = row[i + 1]
        items.append(d)

    return {"total": total, "items": items, "page": page, "page_size": page_size}


@app.get("/api/{entity}/{item_id}")
def get_item(entity: str, item_id: int, db: Session = Depends(get_db)):
    Model = get_model(entity)
    obj = db.query(Model).filter(Model.id == item_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="记录不存在")
    return obj.to_dict()


@app.post("/api/{entity}", status_code=201)
def create_item(entity: str, payload: dict, db: Session = Depends(get_db)):
    Model = get_model(entity)
    cols = get_columns(Model)
    data = {k: v for k, v in payload.items() if k in cols and k not in ("id", "created_at")}
    # 字典类实体（分类/单位）未填写编码时自动生成
    if "code" in cols and not data.get("code"):
        prefix = {"categories": "C", "units": "U", "purchase_orders": "PO"}.get(entity)
        if prefix:
            data["code"] = prefix + "-" + str(db.query(Model).count() + 1).zfill(4)
    enforce_quota(entity, data, db)
    obj = Model(**data)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj.to_dict()


@app.put("/api/{entity}/{item_id}")
def update_item(entity: str, item_id: int, payload: dict, db: Session = Depends(get_db)):
    Model = get_model(entity)
    obj = db.query(Model).filter(Model.id == item_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="记录不存在")
    cols = get_columns(Model)
    # 自动编码（新增时若无 code 则生成）
    old_status = obj.status
    for k, v in payload.items():
        if k in cols and k not in ("id", "created_at"):
            setattr(obj, k, v)
    # 预制计划与库存同步：进入/退出「已完成」时扣减或回补库存
    if entity == "prep":
        new_status = obj.status
        if new_status == "已完成" and old_status != "已完成":
            sync_prep_inventory(obj.materials, obj.store_id, db, True)
        elif old_status == "已完成" and new_status != "已完成":
            sync_prep_inventory(obj.materials, obj.store_id, db, False)
    db.commit()
    db.refresh(obj)
    return obj.to_dict()


@app.delete("/api/{entity}/{item_id}")
def delete_item(entity: str, item_id: int, db: Session = Depends(get_db)):
    Model = get_model(entity)
    obj = db.query(Model).filter(Model.id == item_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(obj)
    db.commit()
    return {"ok": True, "id": item_id}


# ============ 叫货单：门店 → 供应商 → 发货 → 签收 → 库存闭环 ============

@app.post("/api/purchase_orders/{po_id}/approve")
def po_approve(po_id: int, payload: dict = None, db: Session = Depends(get_db)):
    """供应商审核通过叫货单"""
    po = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="叫货单不存在")
    if po.status != "待审核":
        raise HTTPException(status_code=400, detail="当前状态不允许审核")
    po.status = "已通过"
    po.approve_note = (payload or {}).get("note", "")
    db.commit(); db.refresh(po)
    return po.to_dict()


@app.post("/api/purchase_orders/{po_id}/reject")
def po_reject(po_id: int, payload: dict = None, db: Session = Depends(get_db)):
    """供应商审核拒绝叫货单"""
    po = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="叫货单不存在")
    if po.status != "待审核":
        raise HTTPException(status_code=400, detail="当前状态不允许审核")
    note = (payload or {}).get("note", "不符合要求")
    po.status = "已拒绝"
    po.approve_note = note
    db.commit(); db.refresh(po)
    return po.to_dict()


@app.post("/api/purchase_orders/{po_id}/ship")
def po_ship(po_id: int, payload: dict, db: Session = Depends(get_db)):
    """供应商发货：扣减总仓库存 + 生成物流记录"""
    po = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="叫货单不存在")
    if po.status != "已通过":
        raise HTTPException(status_code=400, detail="当前状态不允许发货")

    # 解析叫货明细
    try:
        items = json.loads(po.items or "[]")
    except Exception:
        raise HTTPException(status_code=400, detail="叫货明细数据异常")
    if not items:
        raise HTTPException(status_code=400, detail="叫货明细为空")

    # 逐项扣减总仓库存（store_id=NULL）
    shortages = []
    for it in items:
        iid = it.get("i")
        q = float(it.get("q") or 0)
        if not iid or q <= 0:
            continue
        inv = db.query(models.Inventory).filter(
            models.Inventory.ingredient_id == iid,
            models.Inventory.store_id == None
        ).first()
        if not inv:
            shortages.append(f"{it.get('n','配料#'+str(iid))} 无总仓库存记录")
            continue
        if (inv.current_stock or 0) < q:
            shortages.append(f"{it.get('n','配料#'+str(iid))} 库存不足（需要{q}，当前{inv.current_stock}）")
            continue
        inv.current_stock -= q
        inv.last_out = _today()
        s = inv.safety_stock or 0
        if inv.current_stock <= 0:
            inv.status = "库存不足"
        elif inv.current_stock < s:
            inv.status = "低于安全库存"
        else:
            inv.status = "正常"
    if shortages:
        db.rollback()
        raise HTTPException(status_code=400, detail="；".join(shortages))

    # 生成物流记录
    log_code = "WL" + _today().replace("-", "") + str(po_id).zfill(4)
    log = models.Logistics(
        code=log_code,
        supplier_id=po.supplier_id,
        store_id=po.store_id,
        warehouse=payload.get("warehouse", ""),
        details=", ".join(f"{it.get('n','')}{it.get('q','')}{it.get('u','')}" for it in items),
        total_weight=payload.get("total_weight", ""),
        logistics_company=payload.get("logistics_company", ""),
        ship_date=_today(),
        eta=payload.get("eta", ""),
        status="运输中",
        purchase_order_id=po.id,
    )
    db.add(log); db.flush()
    po.status = "已发货"
    po.logistics_id = log.id
    db.commit(); db.refresh(po)
    return po.to_dict()


@app.post("/api/purchase_orders/{po_id}/receive")
def po_receive(po_id: int, db: Session = Depends(get_db)):
    """门店签收：增加门店库存 + 更新物流状态"""
    po = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="叫货单不存在")
    if po.status != "已发货":
        raise HTTPException(status_code=400, detail="当前状态不允许签收")

    # 解析叫货明细
    try:
        items = json.loads(po.items or "[]")
    except Exception:
        raise HTTPException(status_code=400, detail="叫货明细数据异常")

    # 逐项增加门店库存
    for it in items:
        iid = it.get("i")
        q = float(it.get("q") or 0)
        if not iid or q <= 0:
            continue
        # 找门店库存（store_id = po.store_id）
        inv = db.query(models.Inventory).filter(
            models.Inventory.ingredient_id == iid,
            models.Inventory.store_id == po.store_id
        ).first()
        if not inv:
            # 门店首次入库：创建一条
            inv = models.Inventory(
                ingredient_id=iid, store_id=po.store_id,
                current_stock=0, safety_stock=50,
                location="", status="正常"
            )
            db.add(inv); db.flush()
        inv.current_stock += q
        inv.last_in = _today()
        s = inv.safety_stock or 0
        if inv.current_stock < s:
            inv.status = "低于安全库存"
        else:
            inv.status = "正常"

    # 更新物流状态
    if po.logistics_id:
        log = db.query(models.Logistics).filter(models.Logistics.id == po.logistics_id).first()
        if log:
            log.status = "已签收"
            log.actual_arrival = _today()
    po.status = "已签收"
    db.commit(); db.refresh(po)
    return po.to_dict()


@app.get("/health")
def health():
    return {"status": "ok", "entities": list(MODEL_REGISTRY.keys())}


@app.get("/org-tree")
def org_tree(db: Session = Depends(get_db), x_current_user: str = Header(None, alias="x-current-user")):
    """返回当前用户可见的组织层级树（非平台用户只能看到自己及下游）"""
    acct, user_entity, user_obj = resolve_current_user(x_current_user, db)
    # 平台用户：全量
    if not acct or acct.role == "platform" or not user_obj:
        platforms = db.query(models.Platform).all()
        tree = _build_full_tree(platforms, db)
        return tree
    # 非平台用户：从自己出发递归构建子树
    if acct.role == "supplier":
        sup = db.query(models.Supplier).filter(models.Supplier.id == user_obj.id).first()
        if sup:
            return [_build_supplier_node(sup, db)]
    if acct.role == "brand":
        brand = db.query(models.Brand).filter(models.Brand.id == user_obj.id).first()
        if brand:
            return [_build_brand_node(brand, db)]
    if acct.role == "store":
        store = db.query(models.Store).filter(models.Store.id == user_obj.id).first()
        if store:
            return [_build_store_node(store, db)]
    if acct.role == "staff":
        staff = db.query(models.Staff).filter(models.Staff.id == user_obj.id).first()
        if staff:
            return [{"id": staff.id, "name": staff.name, "type": "店员"}]
    return []


def _build_full_tree(platforms, db):
    tree = []
    for p in platforms:
        p_node = {"id": p.id, "name": p.name, "type": "平台服务商",
                  "quota": p.max_suppliers, "children": []}
        suppliers = db.query(models.Supplier).filter(models.Supplier.platform_id == p.id).all()
        for s in suppliers:
            p_node["children"].append(_build_supplier_node(s, db))
        tree.append(p_node)
    return tree


def _build_supplier_node(s, db):
    node = {"id": s.id, "name": s.name, "type": "供应商", "quota": s.max_brands, "children": []}
    brands = db.query(models.Brand).filter(models.Brand.supplier_id == s.id).all()
    for b in brands:
        node["children"].append(_build_brand_node(b, db))
    return node


def _build_brand_node(b, db):
    node = {"id": b.id, "name": b.name, "type": "品牌商", "quota": b.max_stores, "children": []}
    stores = db.query(models.Store).filter(models.Store.brand_id == b.id).all()
    for st in stores:
        node["children"].append(_build_store_node(st, db))
    return node


def _build_store_node(st, db):
    staffs = db.query(models.Staff).filter(models.Staff.store_id == st.id).all()
    return {
        "id": st.id, "name": st.name, "type": "门店", "quota": st.max_staff,
        "children": [{"id": x.id, "name": x.name, "type": "店员"} for x in staffs],
    }


# ============ 统一账号与审批流 ============
# ROLE_TO_ENTITY 已在权限模块定义（行152），此处复用
ROLE_LABEL = {"platform": "平台服务商", "supplier": "供应商", "brand": "品牌商",
              "store": "门店", "staff": "店员"}
CODE_PREFIX = {"suppliers": "SUP", "brands": "BR", "stores": "ST", "staff": "EMP"}
# 申请表单中排除的账号/系统字段（这些由账号体系自动接管）
APPLY_EXCLUDE = {"username", "password", "account_status", "registered_at", "code"}


def _today():
    from datetime import datetime as _dt
    return _dt.now().strftime("%Y-%m-%d")


@app.post("/auth/register")
def register(payload: dict, db: Session = Depends(get_db)):
    """基础注册：仅填写登录账号/手机号/密码，不绑定任何层级（status=待激活）"""
    username = (payload.get("username") or "").strip()
    password = (payload.get("password") or "").strip()
    phone = (payload.get("phone") or "").strip()
    if not username or not password:
        raise HTTPException(status_code=400, detail="登录账号与密码均不能为空")
    if db.query(models.Account).filter(models.Account.username == username).first():
        raise HTTPException(status_code=400, detail="该登录账号已被注册，请更换")
    acct = models.Account(
        username=username, password=password or "123456", phone=phone,
        role="", status="待激活", registered_at=_today(),
    )
    db.add(acct); db.commit(); db.refresh(acct)
    return acct.to_dict()


@app.post("/auth/login")
def login(payload: dict, db: Session = Depends(get_db)):
    """登录：校验账号密码，返回账号信息（前端按 status 决定跳转去向）"""
    username = (payload.get("username") or "").strip()
    password = (payload.get("password") or "").strip()
    acct = db.query(models.Account).filter(models.Account.username == username).first()
    if not acct or acct.password != password:
        raise HTTPException(status_code=401, detail="登录账号或密码错误")
    return acct.to_dict()


@app.get("/auth/accounts")
def list_accounts(request: Request, db: Session = Depends(get_db)):
    """列出账号（供鸣智后台审批使用），可按 status 过滤，并联动关联实体名称"""
    params = dict(request.query_params)
    status = params.get("status")
    q = db.query(models.Account)
    if status:
        q = q.filter(models.Account.status == status)
    rows = q.order_by(models.Account.id).all()
    items = []
    for a in rows:
        d = a.to_dict()
        d["role_label"] = ROLE_LABEL.get(a.role, "未申请层级")
        d["related_name"] = ""
        if a.role in ROLE_TO_ENTITY and a.related_id:
            E = MODEL_REGISTRY[ROLE_TO_ENTITY[a.role]]
            obj = db.query(E).filter(E.id == a.related_id).first()
            if obj:
                d["related_name"] = obj.name
        items.append(d)
    return {"total": len(items), "items": items, "page": 1, "page_size": len(items)}


@app.get("/auth/accounts/{acct_id}")
def get_account(acct_id: int, db: Session = Depends(get_db)):
    acct = db.query(models.Account).filter(models.Account.id == acct_id).first()
    if not acct:
        raise HTTPException(status_code=404, detail="账号不存在")
    return acct.to_dict()


@app.post("/auth/apply")
def apply(payload: dict, db: Session = Depends(get_db)):
    """登录后申请层级：创建对应实体并挂接账号，账号进入待审批状态"""
    acct_id = payload.get("account_id")
    role = (payload.get("role") or "").strip()
    if role not in ROLE_TO_ENTITY:
        raise HTTPException(status_code=400, detail="请选择有效的层级（供应商/品牌商/门店/店员）")
    acct = db.query(models.Account).filter(models.Account.id == acct_id).first()
    if not acct:
        raise HTTPException(status_code=404, detail="账号不存在，请重新登录")
    if acct.role:
        raise HTTPException(status_code=400, detail="您已提交过层级申请，无需重复提交")
    entity = ROLE_TO_ENTITY[role]
    Model = MODEL_REGISTRY[entity]
    cols = get_columns(Model)
    data = {}
    for k in cols:
        if k in APPLY_EXCLUDE or k in ("id", "created_at"):
            continue
        v = payload.get(k)
        if v in (None, ""):
            continue
        col_type = Model.__table__.c[k].type
        if isinstance(col_type, Integer):
            try:
                v = int(v)
            except (ValueError, TypeError):
                continue
        elif isinstance(col_type, Float):
            try:
                v = float(v)
            except (ValueError, TypeError):
                continue
        data[k] = v
    # 必填校验：名称不可为空
    if not data.get("name"):
        raise HTTPException(status_code=400, detail="请填写名称后再提交申请")
    # 供应商申请时自动挂接鸣智平台
    if entity == "suppliers":
        plat = db.query(models.Platform).order_by(models.Platform.id).first()
        data["platform_id"] = plat.id if plat else None
    enforce_quota(entity, data, db)
    obj = MODEL_REGISTRY[entity](**data)
    # 账号字段回写实体，状态置为待激活（待平台审批后转正）
    obj.username = acct.username
    obj.password = acct.password
    obj.account_status = "待激活"
    obj.registered_at = acct.registered_at
    db.add(obj); db.commit(); db.refresh(obj)
    # 自动生成编码
    if not obj.code:
        obj.code = CODE_PREFIX[entity] + "-" + str(obj.id).zfill(3)
        db.commit()
    # 账号进入待审批
    acct.role = role
    acct.related_id = obj.id
    acct.status = "待审批"
    db.commit(); db.refresh(acct)
    return acct.to_dict()


def _switch_account_status(acct_id: int, approve: bool, note: str, db):
    acct = db.query(models.Account).filter(models.Account.id == acct_id).first()
    if not acct:
        raise HTTPException(status_code=404, detail="账号不存在")
    if acct.role not in ROLE_TO_ENTITY or not acct.related_id:
        raise HTTPException(status_code=400, detail="该账号尚未提交层级申请")
    acct.status = "正常" if approve else "已拒绝"
    acct.approve_note = note or ""
    if approve:
        E = MODEL_REGISTRY[ROLE_TO_ENTITY[acct.role]]
        obj = db.query(E).filter(E.id == acct.related_id).first()
        if obj:
            obj.account_status = "正常"
            obj.username = acct.username
            obj.password = acct.password
            db.commit()
    db.commit(); db.refresh(acct)
    return acct.to_dict()


@app.post("/auth/accounts/{acct_id}/approve")
def approve_account(acct_id: int, payload: dict = None, db: Session = Depends(get_db)):
    """鸣智后台审批通过：账号转正，关联实体启用"""
    note = (payload or {}).get("approve_note", "")
    return _switch_account_status(acct_id, True, note, db)


@app.post("/auth/accounts/{acct_id}/reject")
def reject_account(acct_id: int, payload: dict = None, db: Session = Depends(get_db)):
    """鸣智后台审批驳回：账号置为已拒绝"""
    note = (payload or {}).get("approve_note", "")
    return _switch_account_status(acct_id, False, note, db)


@app.post("/auth/change-pwd")
def change_password(payload: dict, db: Session = Depends(get_db)):
    """修改当前账号密码"""
    acct_id = payload.get("account_id")
    new_pwd = payload.get("new_password", "").strip()
    if not new_pwd or len(new_pwd) < 6:
        raise HTTPException(status_code=400, detail="密码至少 6 位")
    acct = db.query(models.Account).filter(models.Account.id == acct_id).first()
    if not acct:
        raise HTTPException(status_code=404, detail="账号不存在")
    # 同步更新关联实体
    if acct.role in ROLE_TO_ENTITY and acct.related_id:
        E = MODEL_REGISTRY[ROLE_TO_ENTITY[acct.role]]
        obj = db.query(E).filter(E.id == acct.related_id).first()
        if obj:
            obj.password = new_pwd
    acct.password = new_pwd
    db.commit()
    return {"detail": "密码修改成功"}


@app.post("/auth/accounts/{acct_id}/reset")
def reset_account(acct_id: int, db: Session = Depends(get_db)):
    """驳回后重新申请：清除层级与关联实体，回到待激活状态"""
    acct = db.query(models.Account).filter(models.Account.id == acct_id).first()
    if not acct:
        raise HTTPException(status_code=404, detail="账号不存在")
    if acct.role in ROLE_TO_ENTITY and acct.related_id:
        E = MODEL_REGISTRY[ROLE_TO_ENTITY[acct.role]]
        obj = db.query(E).filter(E.id == acct.related_id).first()
        if obj:
            db.delete(obj); db.commit()
    acct.role = ""
    acct.related_id = 0
    acct.status = "待激活"
    acct.level = ""
    acct.approve_note = ""
    db.commit(); db.refresh(acct)
    return acct.to_dict()


# ============ 种子数据 ============
def sync_accounts(db):
    """为已存在的各层级记录补齐统一登录账号（幂等，仅补充缺失账号）。"""
    def _sync(entities, role):
        for e in entities:
            if not e.username:
                continue
            if db.query(models.Account).filter(models.Account.username == e.username).first():
                continue
            db.add(models.Account(
                username=e.username, password=e.password or "123456", phone=e.phone or "",
                role=role, status="正常", related_id=e.id, registered_at=e.registered_at or _today(),
            ))
    platforms = db.query(models.Platform).all()
    _sync(db.query(models.Supplier).all(), "supplier")
    _sync(db.query(models.Brand).all(), "brand")
    _sync(db.query(models.Store).all(), "store")
    _sync(db.query(models.Staff).all(), "staff")
    # 鸣智科技（平台服务商）顶层管理员账号
    if platforms and not db.query(models.Account).filter(models.Account.username == "mingzhi").first():
        db.add(models.Account(
            username="mingzhi", password="123456", phone=platforms[0].phone,
            role="platform", status="正常", related_id=platforms[0].id,
            level="顶级服务商", registered_at=_today(),
        ))
    db.commit()


def seed_dictionaries(db):
    """幂等写入字典数据：配料三级分类 + 常用单位及换算（仅当表为空时）。"""
    if db.query(models.Category).count() == 0:
        # 一级分类
        c1_raw = models.Category(code="C-1001", name="原料", level=1, parent_id=0, sort=1, status="启用")
        c1_aux = models.Category(code="C-1002", name="辅料", level=1, parent_id=0, sort=2, status="启用")
        c1_pkg = models.Category(code="C-1003", name="包装", level=1, parent_id=0, sort=3, status="启用")
        db.add_all([c1_raw, c1_aux, c1_pkg]); db.flush()
        # 二级分类
        c2_tea = models.Category(code="C-2001", name="茶基底", level=2, parent_id=c1_raw.id, sort=1)
        c2_dairy = models.Category(code="C-2002", name="乳基底", level=2, parent_id=c1_raw.id, sort=2)
        c2_aux = models.Category(code="C-2003", name="小料辅料", level=2, parent_id=c1_aux.id, sort=1)
        c2_pkg = models.Category(code="C-2004", name="包装物", level=2, parent_id=c1_pkg.id, sort=1)
        db.add_all([c2_tea, c2_dairy, c2_aux, c2_pkg]); db.flush()
        # 三级分类（叶子，名称与配料 category 一致）
        leaves = [
            ("C-3001", "茶叶类", c2_tea.id), ("C-3002", "奶制品类", c2_dairy.id),
            ("C-3003", "糖浆类", c2_aux.id), ("C-3004", "小料类", c2_aux.id),
            ("C-3005", "果酱类", c2_aux.id), ("C-3006", "水果类", c2_aux.id),
            ("C-3007", "包装材料", c2_pkg.id),
        ]
        for code, name, pid in leaves:
            db.add(models.Category(code=code, name=name, level=3, parent_id=pid, sort=1, status="启用"))
        db.commit()

    if db.query(models.Unit).count() == 0:
        units = [
            models.Unit(code="U-0001", name="千克", symbol="kg", group="重量", base_unit="克", factor=1000),
            models.Unit(code="U-0002", name="克", symbol="g", group="重量", base_unit="克", factor=1),
            models.Unit(code="U-0003", name="毫克", symbol="mg", group="重量", base_unit="克", factor=0.001),
            models.Unit(code="U-0004", name="升", symbol="L", group="容量", base_unit="毫升", factor=1000),
            models.Unit(code="U-0005", name="毫升", symbol="ml", group="容量", base_unit="毫升", factor=1),
            models.Unit(code="U-0006", name="箱", symbol="箱", group="计数", base_unit="袋", factor=24),
            models.Unit(code="U-0007", name="袋", symbol="袋", group="计数", base_unit="袋", factor=1),
            models.Unit(code="U-0008", name="瓶", symbol="瓶", group="计数", base_unit="瓶", factor=1),
            models.Unit(code="U-0009", name="杯", symbol="杯", group="计数", base_unit="杯", factor=1),
        ]
        db.add_all(units); db.commit()


def seed():
    db = next(get_db())
    try:
        seed_dictionaries(db)
        if db.query(models.Supplier).count() > 0:
            sync_accounts(db)
            return
        platforms = [
            models.Platform(code="P-001", name="鸣智科技", contact="平台管理员", phone="400-000-0000",
                            level="顶级服务商", max_suppliers=20, status="运营中"),
        ]
        db.add_all(platforms); db.commit()
        for p in platforms: db.refresh(p)

        # 供应商（上级：平台服务商，向下级授权配额）
        suppliers = [
            models.Supplier(code="SUP-001", name="茗源茶业有限公司", type="贸易商", platform_id=platforms[0].id,
                            contact="陈志远", phone="138****6789", category="茶叶类", level="战略合作",
                            supply_cycle="7天", cooperation_start="2023-01", total_amount=856000, max_brands=3,
                            username="mingyuan", password="123456", account_status="正常", registered_at="2023-01-10", status="合作中"),
            models.Supplier(code="SUP-002", name="甜蜜源糖业", type="制造商", platform_id=platforms[0].id,
                            contact="林小红", phone="139****8901", category="糖浆类", level="战略合作",
                            supply_cycle="5天", cooperation_start="2023-03", total_amount=620000, max_brands=3,
                            username="tianmi", password="123456", account_status="正常", registered_at="2023-03-12", status="合作中"),
            models.Supplier(code="SUP-003", name="光明乳业股份有限公司", type="制造商", platform_id=platforms[0].id,
                            contact="黄大明", phone="137****2345", category="奶制品类", level="核心供应商",
                            supply_cycle="3天", cooperation_start="2022-06", total_amount=1240000, max_brands=2,
                            username="guangming", password="123456", account_status="正常", registered_at="2022-06-01", status="合作中"),
            models.Supplier(code="SUP-004", name="珍珠大王食品", type="贸易商", platform_id=platforms[0].id,
                            contact="吴小军", phone="136****5678", category="小料类", level="一般供应商",
                            supply_cycle="10天", cooperation_start="2024-01", total_amount=180000, max_brands=2,
                            username="zhenzhu", password="123456", account_status="正常", registered_at="2024-01-15", status="合作中"),
            models.Supplier(code="SUP-005", name="鲜果汇供应链", type="贸易商", platform_id=platforms[0].id,
                            contact="周美玲", phone="135****9012", category="水果类", level="核心供应商",
                            supply_cycle="2天", cooperation_start="2023-08", total_amount=420000, max_brands=2,
                            username="xianguo", password="123456", account_status="正常", registered_at="2023-08-20", status="合作中"),
            models.Supplier(code="SUP-006", name="安佳乳品（中国）", type="制造商", platform_id=platforms[0].id,
                            contact="郑国强", phone="133****3456", category="奶制品类", level="一般供应商",
                            supply_cycle="7天", cooperation_start="2024-03", total_amount=95000, max_brands=1,
                            username="anjia", password="123456", account_status="停用", registered_at="2024-03-05", status="暂停合作"),
        ]
        db.add_all(suppliers); db.commit()
        for s in suppliers: db.refresh(s)

        # 品牌商（上级：供应商）
        brands = [
            models.Brand(code="BR-001", name="茶颜悦色", company="茶颜悦色餐饮管理有限公司",
                         supplier_id=suppliers[0].id, manager="李建国", phone="138****0001", store_count=156,
                         franchise_mode="直营+加盟", created_at_date="2022-01", max_stores=20, status="运营中",
                         username="chayan", password="123456", account_status="正常", registered_at="2022-01-15"),
            models.Brand(code="BR-002", name="蜜雪冰城", company="蜜雪冰城股份有限公司",
                         supplier_id=suppliers[1].id, manager="张伟", phone="139****0002", store_count=320,
                         franchise_mode="直营+加盟", created_at_date="2021-06", max_stores=30, status="运营中",
                         username="mixue", password="123456", account_status="正常", registered_at="2021-06-20"),
            models.Brand(code="BR-003", name="喜茶", company="深圳美西西餐饮管理有限公司",
                         supplier_id=suppliers[2].id, manager="刘洋", phone="137****0003", store_count=89,
                         franchise_mode="直营", created_at_date="2022-03", max_stores=10, status="运营中",
                         username="xicha", password="123456", account_status="正常", registered_at="2022-03-10"),
        ]
        db.add_all(brands); db.commit()
        for b in brands: db.refresh(b)

        # 门店
        stores = [
            models.Store(code="ST-001", name="朝阳旗舰店", brand_id=brands[0].id, type="旗舰店",
                         manager="王芳", phone="010-****1234", address="北京市朝阳区建国路88号",
                         area="120㎡", staff_count=8, max_staff=15, status="营业中",
                         username="chaoyang", password="123456", account_status="正常", registered_at="2022-02-01"),
            models.Store(code="ST-002", name="海淀中关村店", brand_id=brands[0].id, type="标准店",
                         manager="李红", phone="010-****5678", address="北京市海淀区中关村大街15号",
                         area="80㎡", staff_count=5, max_staff=10, status="营业中",
                         username="haidian", password="123456", account_status="正常", registered_at="2022-04-12"),
            models.Store(code="ST-003", name="西城金融街店", brand_id=brands[1].id, type="标准店",
                         manager="赵强", phone="010-****9012", address="北京市西城区金融街7号",
                         area="60㎡", staff_count=4, max_staff=8, status="营业中",
                         username="xicheng", password="123456", account_status="正常", registered_at="2022-05-20"),
            models.Store(code="ST-004", name="东城王府井店", brand_id=brands[2].id, type="旗舰店",
                         manager="陈雪", phone="010-****3456", address="北京市东城区王府井大街200号",
                         area="150㎡", staff_count=10, max_staff=12, status="装修中",
                         username="dongcheng", password="123456", account_status="待激活", registered_at="2023-09-01"),
        ]
        db.add_all(stores); db.commit()
        for s in stores: db.refresh(s)

        # 配料
        ings = [
            models.Ingredient(code="PL-001", name="茉莉绿茶", category="茶叶类", spec="25kg/袋", unit="kg",
                              supplier_id=suppliers[0].id, cost_price=85.0, safety_stock=200, status="启用"),
            models.Ingredient(code="PL-002", name="锡兰红茶", category="茶叶类", spec="20kg/箱", unit="kg",
                              supplier_id=suppliers[0].id, cost_price=120.0, safety_stock=150, status="启用"),
            models.Ingredient(code="PL-015", name="鲜牛奶", category="奶制品类", spec="1L/瓶", unit="L",
                              supplier_id=suppliers[2].id, cost_price=12.5, safety_stock=300, status="启用"),
            models.Ingredient(code="PL-028", name="黑糖珍珠", category="小料类", spec="2kg/袋", unit="kg",
                              supplier_id=suppliers[3].id, cost_price=18.0, safety_stock=80, status="启用"),
            models.Ingredient(code="PL-042", name="芒果果酱", category="果酱类", spec="3kg/桶", unit="kg",
                              supplier_id=suppliers[4].id, cost_price=45.0, safety_stock=50, status="低库存"),
            models.Ingredient(code="PL-056", name="果葡糖浆", category="糖浆类", spec="25kg/桶", unit="kg",
                              supplier_id=suppliers[1].id, cost_price=65.0, safety_stock=100, status="启用"),
            models.Ingredient(code="PL-078", name="芝士奶盖粉", category="奶制品类", spec="1kg/袋", unit="kg",
                              supplier_id=suppliers[5].id, cost_price=38.0, safety_stock=60, status="启用"),
            models.Ingredient(code="PL-089", name="椰果", category="小料类", spec="3kg/罐", unit="kg",
                              supplier_id=suppliers[3].id, cost_price=22.0, safety_stock=40, status="停用"),
        ]
        db.add_all(ings); db.commit()
        for i in ings: db.refresh(i)

        # 库存
        invs = [
            models.Inventory(ingredient_id=ings[0].id, current_stock=245, safety_stock=200, status="正常",
                             last_in="06-10", last_out="06-12", location="A-01-03"),
            models.Inventory(ingredient_id=ings[1].id, current_stock=180, safety_stock=150, status="正常",
                             last_in="06-08", last_out="06-11", location="A-01-04"),
            models.Inventory(ingredient_id=ings[2].id, current_stock=120, safety_stock=300, status="库存不足",
                             last_in="06-10", last_out="06-12", location="B-02-01"),
            models.Inventory(ingredient_id=ings[3].id, current_stock=45, safety_stock=80, status="低于安全库存",
                             last_in="06-05", last_out="06-12", location="C-03-02"),
            models.Inventory(ingredient_id=ings[4].id, current_stock=28, safety_stock=50, status="低于安全库存",
                             last_in="06-03", last_out="06-11", location="B-02-05"),
            models.Inventory(ingredient_id=ings[5].id, current_stock=320, safety_stock=100, status="正常",
                             last_in="06-12", last_out="06-10", location="A-01-06"),
        ]
        db.add_all(invs); db.commit()

        # 店员
        staffs = [
            models.Staff(code="EMP-001", name="王芳", role="店长", store_id=stores[0].id, phone="138****1001",
                         hire_date="2023-03-01", permissions="全部权限", qualification="高级", status="在职",
                         username="wangfang", password="123456", account_status="正常", registered_at="2023-03-01"),
            models.Staff(code="EMP-002", name="张明", role="制茶师", store_id=stores[0].id, phone="139****1002",
                         hire_date="2023-06-15", permissions="配料/预制/损耗", qualification="中级", status="在职",
                         username="zhangming", password="123456", account_status="正常", registered_at="2023-06-15"),
            models.Staff(code="EMP-003", name="刘小雪", role="收银员", store_id=stores[0].id, phone="137****1003",
                         hire_date="2024-01-10", permissions="出品/收银", qualification="初级", status="在职",
                         username="liuxiaoxue", password="123456", account_status="正常", registered_at="2024-01-10"),
            models.Staff(code="EMP-004", name="李红", role="店长", store_id=stores[1].id, phone="138****2001",
                         hire_date="2023-04-01", permissions="全部权限", qualification="高级", status="在职",
                         username="lihong", password="123456", account_status="正常", registered_at="2023-04-01"),
            models.Staff(code="EMP-005", name="赵强", role="店长", store_id=stores[2].id, phone="138****3001",
                         hire_date="2023-08-01", permissions="全部权限", qualification="中级", status="在职",
                         username="zhaoqiang", password="123456", account_status="正常", registered_at="2023-08-01"),
            models.Staff(code="EMP-006", name="孙丽丽", role="仓库管理员", store_id=stores[1].id, phone="136****2002",
                         hire_date="2024-03-01", permissions="库存/效期/物流", qualification="中级", status="在职",
                         username="sunlili", password="123456", account_status="正常", registered_at="2024-03-01"),
        ]
        db.add_all(staffs); db.commit()

        # 损耗
        wsts = [
            models.Wastage(code="WS-001", store_id=stores[0].id, ingredient_id=ings[2].id, type="过期损耗",
                           quantity=12, amount=150.0, rate="2.1%", responsible="张明", date="06-12", status="已确认"),
            models.Wastage(code="WS-002", store_id=stores[1].id, ingredient_id=ings[3].id, type="制作损耗",
                           quantity=3, amount=54.0, rate="1.8%", responsible="李红", date="06-11", status="已确认"),
            models.Wastage(code="WS-003", store_id=stores[0].id, ingredient_id=ings[4].id, type="变质损耗",
                           quantity=5, amount=85.0, rate="3.5%", responsible="王芳", date="06-10", status="待审核"),
            models.Wastage(code="WS-004", store_id=stores[2].id, ingredient_id=ings[0].id, type="洒漏损耗",
                           quantity=2, amount=6.8, rate="0.4%", responsible="赵强", date="06-09", status="已处理"),
        ]
        db.add_all(wsts); db.commit()

        # 预制（所需配料结构化，仅取自库存管理中已登记的配料）
        preps = [
            models.Prep(code="YZ-001", name="茶底-茉莉绿茶",
                        materials=json.dumps([{"i": ings[0].id, "n": "茉莉绿茶", "q": 5},
                                              {"i": ings[1].id, "n": "锡兰红茶", "q": 3}], ensure_ascii=False),
                        quantity="50 L", duration="30 min", store_id=stores[0].id, plan_time="06:00",
                        status="已完成"),
            models.Prep(code="YZ-002", name="珍珠-黑糖",
                        materials=json.dumps([{"i": ings[3].id, "n": "黑糖珍珠", "q": 8}], ensure_ascii=False),
                        quantity="20 kg", duration="45 min", store_id=stores[0].id, plan_time="07:00",
                        status="进行中"),
            models.Prep(code="YZ-003", name="芝士奶盖",
                        materials=json.dumps([{"i": ings[2].id, "n": "鲜牛奶", "q": 6},
                                              {"i": ings[5].id, "n": "果葡糖浆", "q": 2}], ensure_ascii=False),
                        quantity="15 L", duration="20 min", store_id=stores[1].id, plan_time="08:00",
                        status="待执行"),
            models.Prep(code="YZ-004", name="芒果果肉",
                        materials=json.dumps([{"i": ings[4].id, "n": "芒果果酱", "q": 4}], ensure_ascii=False),
                        quantity="10 kg", duration="40 min", store_id=stores[2].id, plan_time="07:30",
                        status="待执行"),
        ]
        db.add_all(preps); db.commit()

        # 效期（批次）
        exps = [
            models.Expiry(batch_no="BT20250215", ingredient_id=ings[2].id, production_date="2025-02-15",
                          shelf_life="120天", expiry_date="2025-06-15", batch_qty=200, remaining_qty=0,
                          location="B-02-01", status="已过期"),
            models.Expiry(batch_no="BT20250301", ingredient_id=ings[0].id, production_date="2025-03-01",
                          shelf_life="180天", expiry_date="2025-06-15", batch_qty=100, remaining_qty=15,
                          location="A-01-03", status="紧急"),
            models.Expiry(batch_no="BT20250312", ingredient_id=ings[2].id, production_date="2025-03-12",
                          shelf_life="120天", expiry_date="2025-06-17", batch_qty=300, remaining_qty=80,
                          location="B-02-01", status="警告"),
            models.Expiry(batch_no="BT20250401", ingredient_id=ings[3].id, production_date="2025-04-01",
                          shelf_life="90天", expiry_date="2025-06-30", batch_qty=60, remaining_qty=22,
                          location="C-03-02", status="关注"),
            models.Expiry(batch_no="BT20250510", ingredient_id=ings[5].id, production_date="2025-05-10",
                          shelf_life="365天", expiry_date="2026-05-10", batch_qty=500, remaining_qty=320,
                          location="A-01-06", status="正常"),
        ]
        db.add_all(exps); db.commit()

        # 物流
        logs = [
            models.Logistics(code="WL20250601001", supplier_id=suppliers[0].id, warehouse="杭州总仓",
                             store_id=stores[0].id, details="乌龙茶200kg", total_weight="210 kg",
                             logistics_company="顺丰物流", ship_date="06-12", eta="06-15",
                             actual_arrival="", status="运输中"),
            models.Logistics(code="WL20250601002", supplier_id=suppliers[1].id, warehouse="广州仓",
                             store_id=stores[1].id, details="果葡糖浆500kg", total_weight="520 kg",
                             logistics_company="德邦物流", ship_date="06-12", eta="06-14",
                             actual_arrival="", status="运输中"),
            models.Logistics(code="WL20250601003", supplier_id=suppliers[4].id, warehouse="海南仓",
                             store_id=stores[0].id, details="芒果100kg", total_weight="105 kg",
                             logistics_company="顺丰冷链", ship_date="06-11", eta="06-13",
                             actual_arrival="06-13", status="已签收"),
            models.Logistics(code="WL20250601004", supplier_id=suppliers[3].id, warehouse="厦门仓",
                             store_id=stores[2].id, details="黑糖珍珠50kg", total_weight="52 kg",
                             logistics_company="中通快递", ship_date="06-13", eta="06-16",
                             actual_arrival="", status="待发货"),
        ]
        db.add_all(logs); db.commit()

        # 同步账号体系：为各层级已有记录创建统一登录账号（演示用，默认已审批转正）
        sync_accounts(db)
        print("[seed] 示例数据写入完成")
    finally:
        db.close()


migrate()
seed()

# ============ 静态前端 ============
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "..", "static")

class NoCacheStaticFiles(StaticFiles):
    """自定义静态文件服务：对 JS/CSS/HTML 强制禁止浏览器缓存"""
    async def get_response(self, path: str, scope):
        response = await super().get_response(path, scope)
        if path.endswith(('.js', '.css', '.html')) or scope.get('path', '/').rstrip('/') == '':
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
        return response

if os.path.isdir(STATIC_DIR):
    app.mount("/", NoCacheStaticFiles(directory=STATIC_DIR, html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    print("🧋 奶茶原材料管理系统 v3.0")
    print("🔗 本地访问: http://127.0.0.1:8000")
    print("🌐 若开通隧道，可用 0.0.0.0 访问")
    uvicorn.run(app, host="0.0.0.0", port=8000)
